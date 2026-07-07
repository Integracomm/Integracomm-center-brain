"""Cancelamentos → grw_cancelamentos (fonte = planilhas do time; ClickUp de
cancelados é mal preenchido — decisão do Otávio 07/07/26).

Dois arquivos Google Sheets (PRIVADOS — export público dá 401):
  * "Saídas de Clientes" (dez/25–maio/26): abas "CANCELAMENTOS - <MÊS>" com
    Cliente/Início/Data Saída/Meses/Valor/Plano/Equipe/Motivo Informado; abas
    "TÉRMINOS CONTRATOS" = fim de contrato dos planos START (semântica própria).
  * "Bonificação Squads | 2026" (maio/26+): abas "Saídas <MÊS>" com dois blocos
    lado a lado — CLIENTES EM TRATATIVA (pipeline de retenção, c/ Situação) e
    CLIENTES FORMALIZADOS (saída consumada, c/ Data final/Plano/Meses/Valor).

Estratégia: tenta baixar o xlsx via export público (funciona se o Otávio ligar
"qualquer pessoa com o link – leitor"); sem acesso, usa a cópia em data/
(atualizada manualmente/via Claude). Parsing por NOME de cabeçalho (layout
desliza entre meses). Recarga TOTAL a cada sync (tabela pequena) — idempotente.
Dedup maio/26 (existe nos dois arquivos): fica o registro COM motivo.
"""
from __future__ import annotations

import datetime as dt
import io
import re
import unicodedata
from pathlib import Path
from typing import Any

import httpx
import openpyxl

FILE_SAIDAS = "1KiLRnTvC14wM-VGmi7X6vZv2NKi07WHlBaJjzqIHLhk"      # Saídas de Clientes
FILE_SQUADS = "1rIvOWWlZVJMzouHEELvJD-UyEaSw7-UiXBVaTQzGIqc"      # Bonificação Squads
_DATA = Path(__file__).resolve().parents[3] / "data"
_LOCAL = {FILE_SAIDAS: _DATA / "canc_saidas_clientes.xlsx",
          FILE_SQUADS: _DATA / "canc_bonificacao_squads.xlsx"}

_MES = {"janeiro": 1, "fevereiro": 2, "março": 3, "marco": 3, "abril": 4, "maio": 5,
        "junho": 6, "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10,
        "novembro": 11, "dezembro": 12}

_DDL = """
CREATE TABLE IF NOT EXISTS grw_cancelamentos (
    id          SERIAL PRIMARY KEY,
    tipo        TEXT NOT NULL,      -- cancelamento|termino|tratativa
    fonte       TEXT NOT NULL,      -- saidas|squads
    mes         DATE NOT NULL,      -- 1º dia do mês da aba
    cliente     TEXT NOT NULL,
    data_inicio DATE,
    data_saida  DATE,
    meses       NUMERIC,            -- tempo de casa
    valor       NUMERIC,            -- mensalidade (MRR perdido)
    plano       TEXT,
    equipe      TEXT,               -- squad (Bx-Sy) ou time (FÁBIO/ADS)
    gc          TEXT,               -- responsável no squad
    motivo      TEXT,
    situacao    TEXT,               -- tratativas: andamento da retenção
    updated_at  TIMESTAMPTZ NOT NULL DEFAULT now()
);
CREATE INDEX IF NOT EXISTS idx_grw_canc_mes ON grw_cancelamentos(mes, tipo);
"""


def _norm(s: Any) -> str:
    x = unicodedata.normalize("NFD", str(s or "").strip().lower())
    return "".join(c for c in x if unicodedata.category(c) != "Mn")


def _data(v: Any) -> dt.date | None:
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", str(v or ""))
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return dt.date(y if y > 99 else 2000 + y, mo, d)
    return None


def _num(v: Any) -> float | None:
    if isinstance(v, (int, float)):
        return float(v)
    x = re.sub(r"[^\d,\.\-]", "", str(v or ""))
    if not x:
        return None
    try:
        return float(x.replace(".", "").replace(",", ".")) if "," in x else float(x)
    except ValueError:
        return None


def _mes_da_aba(nome: str) -> dt.date | None:
    n = _norm(nome)
    mes = next((v for k, v in _MES.items() if _norm(k) in n), None)
    if mes is None:
        return None
    m_ano = re.search(r"(20)?2(\d)\b", n.replace("2026!", "2026"))
    ano = 2026
    if "dezembro" in n and "25" not in n:  # CANCELAMENTOS - DEZEMBRO = dez/2025
        ano = 2025
    if m_ano and m_ano.group(0) in ("25", "2025"):
        ano = 2025
    return dt.date(ano, mes, 1)


def fetch_workbook(file_id: str) -> openpyxl.Workbook:
    """Export público (se compartilhado); senão a cópia local em data/."""
    try:
        r = httpx.get(f"https://docs.google.com/spreadsheets/d/{file_id}/export",
                      params={"format": "xlsx"}, timeout=60, follow_redirects=True)
        if r.status_code == 200 and r.content[:2] == b"PK":
            return openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    except httpx.HTTPError:
        pass
    local = _LOCAL[file_id]
    if not local.exists():
        raise FileNotFoundError(f"sem acesso público e sem cópia local: {local.name}")
    return openpyxl.load_workbook(local, read_only=True, data_only=True)


def _headers(row: tuple) -> dict[str, int]:
    return {_norm(c): j for j, c in enumerate(row) if c and str(c).strip()}


def _parse_saidas(wb: openpyxl.Workbook) -> list[dict]:
    out: list[dict] = []
    for aba in wb.sheetnames:
        n = _norm(aba)
        mes = _mes_da_aba(aba)
        if mes is None:
            continue
        tipo = ("cancelamento" if "cancelamento" in n else
                "termino" if "termino" in n else None)
        if tipo is None:
            continue
        ws = wb[aba]
        hdr: dict[str, int] | None = None
        for row in ws.iter_rows(values_only=True):
            if hdr is None:
                h = _headers(row)
                if ("cliente" in h or any(k.startswith("seller") for k in h)):
                    hdr = h
                continue
            def col(*nomes):
                for nm in nomes:
                    for k, j in hdr.items():
                        if k.startswith(nm):
                            return row[j] if j < len(row) else None
                return None
            cliente = col("cliente", "seller")
            if not cliente or not str(cliente).strip():
                continue
            if _norm(cliente).startswith(("em andamento", "total")):
                continue
            out.append({
                "tipo": tipo, "fonte": "saidas", "mes": mes,
                "cliente": str(cliente).strip()[:120],
                "data_inicio": _data(col("inicio")),
                "data_saida": _data(col("data saida", "data termino")),
                "meses": _num(col("meses")),
                "valor": _num(col("valor")),
                "plano": (str(col("plano", "planos") or "").strip()[:40] or None),
                "equipe": (str(col("equipe") or "").strip()[:40] or None),
                "gc": (str(col("assessor") or "").strip()[:60] or None),
                "motivo": (str(col("motivo") or "").strip()[:400] or None),
                "situacao": None,
            })
    return out


def _parse_squads(wb: openpyxl.Workbook) -> list[dict]:
    """Abas "Saídas <MÊS>": dois blocos lado a lado, identificados pelos títulos
    CLIENTES EM TRATATIVA / CLIENTES FORMALIZADOS; sub-cabeçalho na linha
    seguinte define as colunas de cada bloco (o layout desliza entre meses)."""
    out: list[dict] = []
    for aba in wb.sheetnames:
        if "saida" not in _norm(aba):
            continue
        mes = _mes_da_aba(aba)
        if mes is None:
            continue
        ws = wb[aba]
        linhas = list(ws.iter_rows(values_only=True))
        if len(linhas) < 3:
            continue
        col_trat = col_form = None
        for j, c in enumerate(linhas[0]):
            nc = _norm(c)
            if "tratativa" in nc:
                col_trat = j
            if "formalizado" in nc:
                col_form = j
        sub = linhas[1]

        def bloco_hdr(ini: int, fim: int) -> dict[str, int]:
            return {_norm(c): j for j, c in enumerate(sub[ini:fim], start=ini)
                    if c and str(c).strip()}

        blocos = []
        if col_trat is not None:
            fim_t = col_form if col_form is not None else len(sub)
            blocos.append(("tratativa", bloco_hdr(col_trat, fim_t)))
        if col_form is not None:
            blocos.append(("cancelamento", bloco_hdr(col_form, len(sub))))
        for row in linhas[2:]:
            for tipo, h in blocos:
                def col(*nomes):
                    for nm in nomes:
                        for k, j in h.items():
                            if k.startswith(nm):
                                return row[j] if j < len(row) else None
                    return None
                cliente = col("empresa")
                if not cliente or not str(cliente).strip():
                    continue
                out.append({
                    "tipo": tipo, "fonte": "squads", "mes": mes,
                    "cliente": str(cliente).strip()[:120],
                    "data_inicio": _data(col("inicio")),
                    "data_saida": _data(col("data final")),
                    "meses": _num(col("meses")),
                    "valor": _num(col("valor")),
                    "plano": (str(col("plano") or "").strip()[:40] or None),
                    "equipe": None,
                    "gc": (str(col("responsavel") or "").strip()[:60] or None),
                    "motivo": None,
                    "situacao": (str(col("situacao", "obs") or "").strip()[:300] or None),
                })
    return out


def parse_all(wbs: dict[str, openpyxl.Workbook]) -> list[dict]:
    regs = _parse_saidas(wbs[FILE_SAIDAS]) + _parse_squads(wbs[FILE_SQUADS])
    # squad das saídas por squad vem na coluna do MEIO ("Responsável" 1 = squad,
    # 2 = GC) — heurística: se gc parece Bx-Sy vira equipe
    for r in regs:
        if r["gc"] and re.match(r"^B\d-S\d", r["gc"]):
            r["equipe"], r["gc"] = r["gc"], None
    # dedup entre arquivos (maio/26 existe nos dois): chave = cliente-norm + mês;
    # preferimos o registro COM motivo (arquivo Saídas de Clientes)
    vistos: dict[tuple, dict] = {}
    finais: list[dict] = []
    for r in regs:
        if r["tipo"] != "cancelamento":
            finais.append(r)
            continue
        k = (_norm(r["cliente"]).split("|")[0].strip()[:40], r["mes"])
        j = vistos.get(k)
        if j is None:
            vistos[k] = r
            finais.append(r)
        elif r.get("motivo") and not j.get("motivo"):
            finais[finais.index(j)] = r
            vistos[k] = r
    return finais


def sync(conn: Any) -> int:
    wbs = {fid: fetch_workbook(fid) for fid in (FILE_SAIDAS, FILE_SQUADS)}
    regs = parse_all(wbs)
    with conn.cursor() as cur:
        cur.execute(_DDL)
        cur.execute("DELETE FROM grw_cancelamentos")
        for r in regs:
            cur.execute(
                """INSERT INTO grw_cancelamentos
                       (tipo, fonte, mes, cliente, data_inicio, data_saida, meses,
                        valor, plano, equipe, gc, motivo, situacao)
                   VALUES (%(tipo)s,%(fonte)s,%(mes)s,%(cliente)s,%(data_inicio)s,
                           %(data_saida)s,%(meses)s,%(valor)s,%(plano)s,%(equipe)s,
                           %(gc)s,%(motivo)s,%(situacao)s)""", r)
    return len(regs)
